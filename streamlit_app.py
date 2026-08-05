import streamlit as st

import requests

import json

import re

import os

import pandas as pd

import plotly.graph_objects as go

import io

import time

import logging

import fcntl

from typing import Tuple, Dict, Any, List

from openpyxl import load_workbook

from openpyxl.styles import PatternFill

from pathlib import Path

# ===============================
# 基础配置与日志（新增：定位中断原因）
# ===============================
# 创建日志配置

logging.basicConfig(

    level=logging.INFO,

    format="%(asctime)s - %(levelname)s - %(message)s",

    handlers=[

        logging.FileHandler("process_log.log", encoding="utf-8"),

        logging.StreamHandler()

    ]

)

logger = logging.getLogger(__name__)

# 页面配置

st.set_page_config(

    page_title="基于大语言模型的汉语隶属度检测划类平台",



    layout="wide",

    initial_sidebar_state="collapsed",

    menu_items=None

)

# ===============================
# 自定义CSS样式（UI优化版）
# ===============================
custom_css = """
<style>
/* ===== 全局基础样式 ===== */
header {visibility: hidden;}
footer {visibility: hidden;}
[data-testid="stSidebar"] {display: none !important;}
.stApp > div:first-child {padding-top: 1rem;}

/* ===== 主背景渐变 ===== */
.stApp {
    background: linear-gradient(135deg, #f5f7fa 0%, #e4e9f2 100%);
}

/* ===== 主容器卡片化 ===== */
.block-container {
    padding-top: 1.5rem;
    padding-bottom: 2rem;
    max-width: 95% !important;
}

/* ===== 标题高亮卡片 ===== */
.title-header-card {
    background: linear-gradient(135deg, #0f2942 0%, #1e4d7b 40%, #2d6cb8 75%, #3d8bd6 100%);
    padding: 2rem 2.5rem;
    border-radius: 20px;
    margin-bottom: 1.5rem;
    box-shadow: 0 12px 40px rgba(15, 41, 66, 0.35);
    position: relative;
    overflow: hidden;
}
.title-header-card::before {
    content: '';
    position: absolute;
    top: -50%;
    right: -10%;
    width: 300px;
    height: 300px;
    background: radial-gradient(circle, rgba(255,255,255,0.1) 0%, transparent 70%);
    border-radius: 50%;
}
.title-header-card h1 {
    color: #ffffff !important;
    font-size: 1.8rem !important;
    font-weight: 700 !important;
    margin: 0 !important;
    padding: 0 !important;
    text-shadow: 0 2px 4px rgba(0,0,0,0.2);
    position: relative;
    z-index: 1;
}
.title-header-card .subtitle {
    color: rgba(255, 255, 255, 0.8) !important;
    font-size: 0.95rem !important;
    margin-top: 0.5rem !important;
    font-style: italic;
    position: relative;
    z-index: 1;
}
.title-header-card .badges {
    margin-top: 1rem;
    display: flex;
    gap: 0.75rem;
    flex-wrap: wrap;
    position: relative;
    z-index: 1;
}
.title-header-card .badge {
    background: rgba(255, 255, 255, 0.15);
    backdrop-filter: blur(10px);
    color: #fff;
    padding: 0.35rem 0.85rem;
    border-radius: 20px;
    font-size: 0.8rem;
    border: 1px solid rgba(255, 255, 255, 0.2);
}

/* ===== 模块高亮卡片 ===== */
.module-card {
    background: #ffffff;
    border-radius: 16px;
    padding: 1.5rem;
    box-shadow: 0 4px 20px rgba(0, 0, 0, 0.08), 0 1px 3px rgba(0, 0, 0, 0.05);
    border: 1px solid rgba(200, 210, 225, 0.5);
    margin-bottom: 1rem;
    transition: all 0.3s ease;
}
.module-card:hover {
    box-shadow: 0 8px 30px rgba(30, 58, 95, 0.12), 0 2px 6px rgba(0, 0, 0, 0.06);
    border-color: rgba(45, 90, 135, 0.2);
}

/* ===== 模型设置区高亮容器 ===== */
.model-settings-card {
    background: linear-gradient(135deg, #fbfcfe 0%, #eef3f9 100%);
    border-radius: 16px;
    padding: 1.25rem 1.5rem;
    border: 1.5px solid rgba(45, 90, 135, 0.2);
    box-shadow: 0 4px 16px rgba(45, 90, 135, 0.1), 0 1px 4px rgba(0, 0, 0, 0.04);
    margin-bottom: 0.5rem;
}
.model-settings-card .section-title {
    margin-top: 0;
    padding-top: 0;
    border-bottom: 2px solid rgba(45, 90, 135, 0.15);
}

/* ===== 连接测试区高亮容器 ===== */
.connection-card {
    background: linear-gradient(135deg, #f5f9ff 0%, #e8f0fe 100%);
    border-radius: 16px;
    padding: 1.25rem 1.5rem;
    border: 1.5px solid rgba(59, 130, 246, 0.25);
    box-shadow: 0 4px 16px rgba(59, 130, 246, 0.12), 0 1px 4px rgba(0, 0, 0, 0.04);
    text-align: center;
    margin-bottom: 0.5rem;
}
.connection-card .section-title {
    margin-top: 0;
    padding-top: 0;
    border-bottom: 2px solid rgba(59, 130, 246, 0.2);
    justify-content: center;
}

/* ===== 结果成功高亮块 ===== */
.result-success-card {
    background: linear-gradient(135deg, #ecfdf5 0%, #d1fae5 100%);
    border-radius: 10px;
    padding: 1rem 1.25rem;
    border-left: 4px solid #10b981;
    margin: 1rem 0;
}

/* ===== 警告高亮块 ===== */
.warning-highlight {
    background: linear-gradient(135deg, #fffbeb 0%, #fef3c7 100%);
    border-radius: 10px;
    padding: 1rem 1.25rem;
    border-left: 4px solid #f59e0b;
    margin: 0.75rem 0;
}

/* ===== 错误高亮块 ===== */
.error-highlight {
    background: linear-gradient(135deg, #fef2f2 0%, #fee2e2 100%);
    border-radius: 10px;
    padding: 1rem 1.25rem;
    border-left: 4px solid #ef4444;
    margin: 0.75rem 0;
}

/* ===== 信息高亮块 ===== */
.info-highlight {
    background: linear-gradient(135deg, #eff6ff 0%, #dbeafe 100%);
    border-radius: 10px;
    padding: 1rem 1.25rem;
    border-left: 4px solid #3b82f6;
    margin: 0.75rem 0;
}

/* ===== 子标题样式 ===== */
.section-title {
    display: flex;
    align-items: center;
    gap: 0.5rem;
    font-size: 1.15rem !important;
    font-weight: 600 !important;
    color: #1e3a5f !important;
    margin-bottom: 1rem !important;
    padding-bottom: 0.5rem;
    border-bottom: 2px solid #e2e8f0;
}
.section-title .icon-dot {
    width: 10px;
    height: 10px;
    border-radius: 50%;
    background: linear-gradient(135deg, #2d6cb8 0%, #3d8bd6 100%);
    box-shadow: 0 2px 6px rgba(45, 108, 184, 0.4);
    flex-shrink: 0;
}

/* ===== 主按钮样式优化 ===== */
.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #1e4d7b 0%, #2d6cb8 50%, #3d8bd6 100%) !important;
    border: none !important;
    border-radius: 12px !important;
    padding: 0.75rem 2rem !important;
    font-weight: 700 !important;
    font-size: 1rem !important;
    color: white !important;
    box-shadow: 0 6px 20px rgba(30, 77, 123, 0.4), 0 2px 6px rgba(0, 0, 0, 0.1) !important;
    transition: all 0.3s ease !important;
    width: 100%;
    letter-spacing: 0.5px;
}
.stButton > button[kind="primary"]:hover {
    background: linear-gradient(135deg, #0f2942 0%, #1e4d7b 50%, #2d6cb8 100%) !important;
    box-shadow: 0 10px 30px rgba(30, 77, 123, 0.5), 0 4px 10px rgba(0, 0, 0, 0.15) !important;
    transform: translateY(-2px);
}
.stButton > button[kind="primary"]:active {
    transform: translateY(0);
    box-shadow: 0 3px 10px rgba(30, 77, 123, 0.3) !important;
}

/* ===== 次要按钮样式 ===== */
.stButton > button[kind="secondary"] {
    background: #ffffff !important;
    border: 2px solid #cbd5e1 !important;
    border-radius: 12px !important;
    padding: 0.6rem 1.5rem !important;
    font-weight: 600 !important;
    color: #475569 !important;
    transition: all 0.3s ease !important;
    box-shadow: 0 2px 8px rgba(0, 0, 0, 0.04);
}
.stButton > button[kind="secondary"]:hover {
    border-color: #2d6cb8 !important;
    color: #1e4d7b !important;
    background: #f5f9ff !important;
    box-shadow: 0 4px 12px rgba(45, 108, 184, 0.15);
    transform: translateY(-1px);
}

/* ===== 输入框样式 ===== */
.stTextInput > div > div > input,
.stTextArea > div > div > textarea,
.stSelectbox > div > div > div {
    border-radius: 12px !important;
    border: 2px solid #e2e8f0 !important;
    transition: all 0.3s ease !important;
    background: #fefefe !important;
}
.stTextInput > div > div > input:focus,
.stTextArea > div > div > textarea:focus,
.stSelectbox > div > div > div:focus {
    border-color: #2d6cb8 !important;
    box-shadow: 0 0 0 4px rgba(45, 108, 184, 0.1) !important;
    background: #ffffff !important;
}

/* ===== 标签页样式优化 ===== */
.stTabs [data-baseweb="tab-list"] {
    gap: 0.5rem;
    background: #f1f5f9;
    padding: 0.5rem;
    border-radius: 12px;
    margin-bottom: 1rem;
}
.stTabs [data-baseweb="tab"] {
    height: 2.5rem;
    border-radius: 8px;
    padding: 0 1.5rem;
    font-weight: 600;
    color: #64748b;
    transition: all 0.3s ease;
}
.stTabs [data-baseweb="tab"]:hover {
    color: #1e4d7b;
    background: rgba(45, 108, 184, 0.08);
}
.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, #1e4d7b 0%, #2d6cb8 100%);
    color: white;
    font-weight: 700;
    box-shadow: 0 2px 8px rgba(30, 77, 123, 0.2);
    border-radius: 8px;
}

/* ===== 表格样式优化 ===== */
.stDataFrame {
    border-radius: 16px;
    overflow: hidden;
    box-shadow: 0 4px 16px rgba(0, 0, 0, 0.06);
    border: 1px solid #e2e8f0;
}
.dataframe {
    font-size: 13px;
}
.dataframe th {
    background: linear-gradient(135deg, #1e4d7b 0%, #2d6cb8 100%) !important;
    color: white !important;
    font-weight: 600 !important;
    padding: 0.85rem 1rem !important;
}
.dataframe td {
    padding: 0.7rem 1rem !important;
}
.dataframe tr:nth-child(even) {
    background: #f8fafc;
}
.dataframe tr:hover {
    background: #eff6ff !important;
}

/* ===== 指标卡片样式 ===== */
.metric-card {
    background: linear-gradient(135deg, #ffffff 0%, #f5f9ff 100%);
    border-radius: 16px;
    padding: 1.25rem 1.5rem;
    border: 1.5px solid rgba(45, 108, 184, 0.15);
    text-align: center;
    box-shadow: 0 4px 12px rgba(45, 108, 184, 0.08);
}
.metric-card .metric-value {
    font-size: 2rem;
    font-weight: 800;
    color: #1e4d7b;
}
.metric-card .metric-label {
    font-size: 0.9rem;
    color: #64748b;
    margin-top: 0.5rem;
    font-weight: 500;
}

/* ===== 进度条样式 ===== */
.stProgress > div > div > div > div {
    background: linear-gradient(90deg, #1e4d7b 0%, #2d6cb8 50%, #3d8bd6 100%) !important;
    border-radius: 10px !important;
    height: 12px !important;
}
.stProgress > div > div > div {
    border-radius: 10px !important;
    background: #e2e8f0 !important;
    height: 12px !important;
}

/* ===== 展开器样式 ===== */
.streamlit-expanderHeader {
    background: #f8fafc;
    border-radius: 12px !important;
    padding: 0.85rem 1.25rem !important;
    font-weight: 600 !important;
    color: #1e4d7b !important;
    border: 1.5px solid #e2e8f0;
    transition: all 0.3s ease;
}
.streamlit-expanderHeader:hover {
    background: #f0f7ff;
    border-color: #2d6cb8;
}
.streamlit-expanderContent {
    background: #ffffff;
    border: 1.5px solid #e2e8f0;
    border-top: none;
    border-radius: 0 0 12px 12px !important;
    padding: 1.25rem !important;
}

/* ===== 代码块样式 ===== */
.stCodeBlock {
    border-radius: 12px !important;
    border: 1.5px solid #e2e8f0 !important;
    box-shadow: 0 2px 8px rgba(0, 0, 0, 0.04) !important;
}

/* ===== 文件上传器样式 ===== */
.stFileUploader {
    background: #ffffff;
    border-radius: 16px;
    padding: 1.5rem;
    border: 2px dashed #cbd5e1;
    transition: all 0.3s ease;
    box-shadow: 0 2px 8px rgba(0, 0, 0, 0.03);
}
.stFileUploader:hover {
    border-color: #2d6cb8;
    background: #f5f9ff;
    box-shadow: 0 4px 12px rgba(45, 108, 184, 0.1);
}

/* ===== 分割线样式 ===== */
hr {
    border: none;
    height: 1px;
    background: linear-gradient(90deg, transparent 0%, #cbd5e1 50%, transparent 100%);
    margin: 1.5rem 0;
}

/* ===== 底部版权 ===== */
.footer-text {
    text-align: center;
    color: #64748b;
    font-size: 0.85rem;
    padding: 1.5rem 0;
    margin-top: 2rem;
    border-top: 1px solid #e5e7eb;
}

/* ===== 状态徽章 ===== */
.status-badge {
    display: inline-block;
    padding: 0.25rem 0.75rem;
    border-radius: 20px;
    font-size: 0.8rem;
    font-weight: 500;
}
.status-badge.success {
    background: #d1fae5;
    color: #065f46;
}
.status-badge.warning {
    background: #fef3c7;
    color: #92400e;
}
.status-badge.error {
    background: #fee2e2;
    color: #991b1b;
}

/* ===== 排名卡片 ===== */
.rank-card {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 0.85rem 1.25rem;
    border-radius: 12px;
    margin-bottom: 0.6rem;
    background: linear-gradient(135deg, #fefefe 0%, #f5f7fa 100%);
    border: 1.5px solid #e2e8f0;
    transition: all 0.3s ease;
    box-shadow: 0 2px 6px rgba(0, 0, 0, 0.03);
}
.rank-card:hover {
    background: linear-gradient(135deg, #f5f9ff 0%, #e8f0fe 100%);
    border-color: #2d6cb8;
    transform: translateX(6px);
    box-shadow: 0 4px 12px rgba(45, 108, 184, 0.12);
}
.rank-card .rank-num {
    width: 32px;
    height: 32px;
    border-radius: 50%;
    background: linear-gradient(135deg, #e2e8f0 0%, #cbd5e1 100%);
    color: #64748b;
    display: flex;
    align-items: center;
    justify-content: center;
    font-weight: 700;
    font-size: 0.9rem;
    box-shadow: 0 2px 4px rgba(0, 0, 0, 0.08);
}
.rank-card.top-1 .rank-num {
    background: linear-gradient(135deg, #fbbf24 0%, #f59e0b 100%);
    color: white;
    box-shadow: 0 3px 8px rgba(251, 191, 36, 0.4);
}
.rank-card.top-2 .rank-num {
    background: linear-gradient(135deg, #9ca3af 0%, #6b7280 100%);
    color: white;
    box-shadow: 0 3px 8px rgba(156, 163, 175, 0.4);
}
.rank-card.top-3 .rank-num {
    background: linear-gradient(135deg, #d97706 0%, #b45309 100%);
    color: white;
    box-shadow: 0 3px 8px rgba(217, 119, 6, 0.4);
}

/* ===== 响应式调整 ===== */
@media (max-width: 768px) {
    .title-header-card {
        padding: 1.5rem 1.25rem;
    }
    .title-header-card h1 {
        font-size: 1.4rem !important;
    }
    .block-container {
        padding-left: 1rem;
        padding-right: 1rem;
    }
}
</style>
"""

st.markdown(custom_css, unsafe_allow_html=True)

# 全局常量（修复：使用绝对路径避免文件路径问题）

BASE_DIR = Path(__file__).parent

BACKUP_FILE = BASE_DIR / "batch_history_log.csv"

PROGRESS_FILE = BASE_DIR / "process_progress.json"  # 新增：保存处理进度

RULE_SETS = {

    "名词": [

        {"name": "N1_可受数量词修饰", "desc": "可以受数量词修饰", "match_score": 10, "mismatch_score": 0},

        {"name": "N2_不能受副词修饰", "desc": "不能受副词修饰", "match_score": 20, "mismatch_score": -20},

        {"name": "N3_可作主宾语", "desc": "可以做典型的主语或宾语", "match_score": 20, "mismatch_score": 0},

        {"name": "N4_可作中心语或作定语", "desc": "可以做中心语受其他名词修饰，或者作定语直接修饰其他名词", "match_score": 10, "mismatch_score": 0},

        {"name": "N5_可后附的字结构", "desc": "可以后附助词'的'构成'的'字结构", "match_score": 10, "mismatch_score": 0},

        {"name": "N6_可后附方位词构处所", "desc": "可以后附方位词构成处所结构", "match_score": 10, "mismatch_score": 0},

        {"name": "N7_不能作谓语核心", "desc": "不能做谓语或谓语核心", "match_score": 10, "mismatch_score": -10},

        {"name": "N8_不能作补语/一般不作状语", "desc": "不能作补语，并且一般不能做状语直接修饰动词性成分", "match_score": 10, "mismatch_score": 0},

    ],

    "动词": [

        {"name": "V1_可受否定'不/没有'修饰", "desc": "可以受否定副词'不'或'没有'修饰", "match_score": 10, "mismatch_score": 0},

        {"name": "V2_可后附/插入时体助词'着/了/过'", "desc": "可以后附或中间插入时体助词'着/了/过'", "match_score": 10, "mismatch_score": 0},

        {"name": "V3_可带真宾语或通过介词引导论元", "desc": "可以带真宾语或通过介词引导论元", "match_score": 20, "mismatch_score": 0},

        {"name": "V4_程度副词与带宾语的关系", "desc": "不能受程度副词'很'修饰，或能同时受'很'修饰并带宾语", "match_score": 10, "mismatch_score": -10},

        {"name": "V5_可有重叠/正反重叠形式", "desc": "可以有'VV, V一V, V了V, V不V, V了没有'等形式", "match_score": 10, "mismatch_score": 0},

        {"name": "V6_可做谓语或谓语核心", "desc": "可以做谓语或谓语核心", "match_score": 10, "mismatch_score": -10},

        {"name": "V7_不能作状语修饰动词性成分", "desc": "不能作状语修饰动词性成分", "match_score": 10, "mismatch_score": 0},

        {"name": "V8_可作'怎么/怎样'提问或'这么/这样/那么'回答", "desc": "可以跟在'怎么/怎样'之后提问或跟在'这么/这样/那么'之后回答", "match_score": 10, "mismatch_score": 0},

        {"name": "V9_不能跟在'多/多么'之后提问或表示感叹", "desc": "不能跟在'多'之后对性质提问，不能跟在'多么'之后表示感叹", "match_score": 10, "mismatch_score": -10},

    ],

    "名动词": [

        {"name": "NV1_可被\"不/没有\"否定且肯定形式-1", "desc": "可以用\"不\"和\"没有\"来否定", "match_score": 10, "mismatch_score": -10},

        {"name": "NV2_可附时体助词或进入\"……了没有\"格式", "desc": "可以后附时体助词\"着、了、过\"", "match_score": 10, "mismatch_score": -10},

        {"name": "NV3_可带真宾语且不受\"很\"修饰", "desc": "可以带真宾语，并且不能受程度副词\"很\"等修饰", "match_score": 10, "mismatch_score": -10},

        {"name": "NV4_有重叠和正反重叠形式", "desc": "可以有\"VV、V一V、V了V、V不V\"等重叠和正反重叠形式", "match_score": 10, "mismatch_score": 0},

        {"name": "NV5_可作多种句法成分且可作形式动词宾语", "desc": "既可以作谓语或谓语核心，又可以作主语或宾语", "match_score": 10, "mismatch_score": -10},

        {"name": "NV6_不能直接作状语", "desc": "不能直接作状语修饰动词性成分", "match_score": 10, "mismatch_score": -10},

        {"name": "NV7_可修饰名词或受名词/数量词修饰", "desc": "可以修饰名词或者受名词修饰，或者可以受数量词修饰", "match_score": 10, "mismatch_score": 0},

        {"name": "NV8_可跟在\"怎么/怎样/这么/这样/那么/那样\"之后", "desc": "可以跟在\"怎么、怎样\"之后提问", "match_score": 10, "mismatch_score": 0},

        {"name": "NV9_不能跟在\"多/多么\"之后", "desc": "不能跟在\"多\"之后对性质的程度进行提问", "match_score": 10, "mismatch_score": -10},

        {"name": "NV10_可后附方位词构成处所结构", "desc": "可以后附方位词构成处所结构", "match_score": 10, "mismatch_score": 0},

    ]

}

# ===============================
# 模型配置
# ===============================

MODEL_CONFIGS = {

    "deepseek": {

        "base_url": "https://api.deepseek.com/v1",

        "endpoint": "/chat/completions",

        "headers": lambda key: {"Authorization": f"Bearer {key}", "Content-Type": "application/json"},

        "payload": lambda model, messages, **kw: {

            "model": model, "messages": messages, "max_tokens": kw.get("max_tokens", 4096), 

            "temperature": kw.get("temperature", 0.0), 

            "stream": True, 

        },

    },

    "openai": {

        "base_url": "https://api.openai.com/v1",

        "endpoint": "/chat/completions",

        "headers": lambda key: {"Authorization": f"Bearer {key}", "Content-Type": "application/json"},

        "payload": lambda model, messages, **kw: {

            "model": model, "messages": messages, "max_tokens": kw.get("max_tokens", 4096), 

            "temperature": kw.get("temperature", 0.0), 

            "stream": True,

        },

    },

    "gemini": {

        # 核心修正：Base URL 到版本号级，Endpoint 为标准 chat 接口

        "base_url": "https://generativelanguage.googleapis.com/v1beta",

        "endpoint": "/chat/completions",

        "headers": lambda key: {"Authorization": f"Bearer {key}", "Content-Type": "application/json"},

        "payload": lambda model, messages, **kw: {

            "model": model, 

            "messages": messages, 

            "max_tokens": kw.get("max_tokens", 4096), 

            "temperature": kw.get("temperature", 0.0), 

            "stream": True,

        },

    },

    "moonshot": {

        "base_url": "https://api.moonshot.cn/v1",

        "endpoint": "/chat/completions",

        "headers": lambda key: {"Authorization": f"Bearer {key}", "Content-Type": "application/json"},

        "payload": lambda model, messages, **kw: {

            "model": model, "messages": messages, "max_tokens": kw.get("max_tokens", 4096), 

            "temperature": kw.get("temperature", 0.0), 

            "stream": True,

        },

    },

    "qwen": {

        "base_url": "https://dashscope.aliyuncs.com/api/v1",

        "endpoint": "/services/aigc/text-generation/generation",

        "headers": lambda key: {

            "Authorization": f"Bearer {key}", 

            "Content-Type": "application/json",

            "X-DashScope-SSE": "enable",

            "Accept": "text/event-stream"

        },

        "payload": lambda model, messages, **kw: {

            "model": model, 

            "input": {"messages": messages}, 

            "parameters": {

                "max_tokens": kw.get("max_tokens", 4096), 

                "temperature": kw.get("temperature", 0.0),

                "result_format": "message",

                "incremental_output": True 

            },

        },

    },

}

# ===============================
# 模型选项（修正 Gemini 模型名称）
# ===============================

MODEL_OPTIONS = {

    "DeepSeek Chat": {

        "provider": "deepseek", 

        "model": "deepseek-chat", 

        "api_key": os.getenv("DEEPSEEK_API_KEY"),

        "env_var": "DEEPSEEK_API_KEY"

    },

    "OpenAI GPT-4o（推荐）": {

        "provider": "openai", 

        "model": "gpt-4o-mini", 

        "api_key": os.getenv("OPENAI_API_KEY"),

        "env_var": "OPENAI_API_KEY"

    },

    "Google Gemini 1.5 Pro": {

        "provider": "gemini", 

        "model": "models/gemini-1.5-pro",  # 关键点：增加 models/ 前缀

        "api_key": os.getenv("GEMINI_API_KEY"),

        "env_var": "GEMINI_API_KEY"

    },

    "Google Gemini 1.5 Flash": {

        "provider": "gemini", 

        "model": "models/gemini-1.5-flash", # 关键点：增加 models/ 前缀

        "api_key": os.getenv("GEMINI_API_KEY"),

        "env_var": "GEMINI_API_KEY"

    },

    "Moonshot（Kimi）": {

        "provider": "moonshot", 

        "model": "moonshot-v1-32k", 

        "api_key": os.getenv("MOONSHOT_API_KEY"),

        "env_var": "MOONSHOT_API_KEY"

    },

    "Qwen（通义千问）": {

        "provider": "qwen", 

        "model": "qwen-max", 

        "api_key": os.getenv("QWEN_API_KEY"),

        "env_var": "QWEN_API_KEY"

    },

}

AVAILABLE_MODEL_OPTIONS = {

    name: info for name, info in MODEL_OPTIONS.items() if info["api_key"]

}

if not AVAILABLE_MODEL_OPTIONS:

    AVAILABLE_MODEL_OPTIONS = MODEL_OPTIONS

# ===============================
# 增强型工具函数（解决中断核心）
# ===============================

def extract_text_from_response(resp_json: Dict[str, Any]) -> str:

    """从不同格式的LLM响应中安全提取文本内容。"""

    if not isinstance(resp_json, dict):

        return ""

    try:

        if "output" in resp_json and "text" in resp_json["output"]:

            return resp_json["output"]["text"]

        if "choices" in resp_json and len(resp_json["choices"]) > 0:

            choice = resp_json["choices"][0]

            if "message" in choice and "content" in choice["message"]:

                return choice["message"]["content"]

        return json.dumps(resp_json, ensure_ascii=False)

    except Exception as e:

        logger.error(f"提取响应文本失败: {e}")

        return json.dumps(resp_json, ensure_ascii=False)

def extract_json_from_text(text: str) -> Tuple[Dict[str, Any], str]:

    """从混合文本中提取并解析JSON对象。"""

    match = re.search(r"(\{.*\})", text.strip(), re.DOTALL)

    if not match:

        return None, text

    json_text = match.group(1).strip()

    try:

        parsed_json = json.loads(json_text)

        return parsed_json, json_text

    except json.JSONDecodeError as e:

        logger.error(f"解析JSON失败: {e}, 原始文本: {json_text[:100]}")

        return None, json_text

def normalize_key(k: str, pos_rules: list) -> str:

    """标准化模型返回的规则名称"""

    if not isinstance(k, str): return None

    k_norm = re.sub(r'[\s_]+', '', k).upper()

    for r in pos_rules:

        r_norm = re.sub(r'[\s_]+', '', r["name"]).upper()

        if r_norm == k_norm:

            return r["name"]

    return None

def map_to_allowed_score(rule: dict, raw_val) -> int:

    """将模型返回值映射为规则得分"""

    match_score, mismatch_score = rule["match_score"], rule["mismatch_score"]

    try:

        if isinstance(raw_val, bool):

            return match_score if raw_val else mismatch_score

        if isinstance(raw_val, str):

            s = raw_val.strip().lower()

            if s in ("yes", "y", "true", "是", "√", "符合"):

                return match_score

            if s in ("no", "n", "false", "否", "×", "不符合"):

                return mismatch_score

        if isinstance(raw_val, (int, float)):

            raw_val_int = int(raw_val)

            if raw_val_int == match_score: return match_score

            if raw_val_int == mismatch_score: return mismatch_score

    except Exception as e:

        logger.error(f"映射得分失败: {e}")

    return mismatch_score

def calculate_membership(scores_all: Dict[str, Dict[str, int]]) -> Dict[str, float]:

    """计算隶属度"""

    membership = {}

    try:

        for pos, scores in scores_all.items():

            total_score = sum(scores.values())

            normalized = total_score / 100

            membership[pos] = max(-1.0, min(1.0, normalized))

    except Exception as e:

        logger.error(f"计算隶属度失败: {e}")

    return membership

def get_top_10_positions(membership: Dict[str, float]) -> List[Tuple[str, float]]:

    """获取隶属度最高的前 10 个词类"""

    try:

        return sorted(membership.items(), key=lambda x: x[1], reverse=True)[:10]

    except Exception as e:

        logger.error(f"排序隶属度失败: {e}")

        return []

def get_history_count(backup_file):

    """获取最新的历史记录数量（实时更新用）"""

    if not os.path.exists(backup_file):

        return 0

    try:

        temp_history = pd.read_csv(backup_file, encoding='utf-8-sig')

        return len(temp_history)

    except Exception as e:

        logger.warning(f"读取历史记录数量失败: {e}")

        return 0

# 新增：文件写入加锁 + 重试（解决文件操作中断）

def safe_write_csv(df, file_path, mode='a', header=False, encoding='utf-8-sig', max_retries=3):

    """安全写入CSV，加文件锁避免冲突，失败自动重试"""

    retry_count = 0

    while retry_count < max_retries:

        try:

            with open(file_path, mode, encoding=encoding) as f:

                fcntl.flock(f, fcntl.LOCK_EX)  # 排他锁

                df.to_csv(f, mode=mode, header=header, index=False)

                fcntl.flock(f, fcntl.LOCK_UN)  # 释放锁

            return True

        except Exception as e:

            retry_count += 1

            logger.warning(f"写入CSV失败（重试{retry_count}/{max_retries}）: {e}")

            time.sleep(1)

    logger.error(f"写入CSV最终失败: {file_path}")

    return False

# 新增：保存/加载处理进度（解决断点续传中断）

def save_process_progress(file_name, current_row, total_rows):

    """保存处理进度"""

    try:

        progress_data = {

            "file_name": file_name,

            "current_row": current_row,

            "total_rows": total_rows,

            "last_update": time.strftime("%Y-%m-%d %H:%M:%S")

        }

        with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:

            json.dump(progress_data, f, ensure_ascii=False, indent=2)

    except Exception as e:

        logger.error(f"保存进度失败: {e}")

def load_process_progress():

    """加载处理进度"""

    if not os.path.exists(PROGRESS_FILE):

        return None

    try:

        with open(PROGRESS_FILE, 'r', encoding='utf-8') as f:

            return json.load(f)

    except Exception as e:

        logger.error(f"加载进度失败: {e}")

        return None

# 新增：清除进度文件

def clear_process_progress():

    """清除进度文件"""

    if os.path.exists(PROGRESS_FILE):

        try:

            os.remove(PROGRESS_FILE)

        except Exception as e:

            logger.error(f"清除进度文件失败: {e}")

# ===============================
# 增强型LLM调用（集成调试与路径修复）
# ===============================

def call_llm_api_cached(_provider, _model, _api_key, messages, max_tokens=4096, temperature=0.0, max_retries=3):

    """封装LLM调用逻辑，彻底解决路径拼接与格式兼容问题"""

    if not _api_key: 

        return False, {"error": "API Key 为空"}, "API Key 未提供"

    if _provider not in MODEL_CONFIGS: 

        return False, {"error": f"未知提供商 {_provider}"}, f"未知提供商 {_provider}"

    cfg = MODEL_CONFIGS[_provider]

    # 显式处理 URL 拼接，避免多余或丢失斜杠

    base_url = cfg['base_url'].rstrip('/')

    endpoint = cfg['endpoint'].lstrip('/')

    url = f"{base_url}/{endpoint}"

    

    headers = cfg["headers"](_api_key)

    payload = cfg["payload"](_model, messages, max_tokens=max_tokens, temperature=temperature)

    streaming_placeholder = st.empty()

    full_content = ""

    error_msg = "未知错误"

    for attempt in range(max_retries):

        try:

            with requests.post(url, headers=headers, json=payload, stream=True, timeout=120) as response:

                # 状态码非 200 处理

                if response.status_code != 200:

                    status_code = response.status_code

                    try:

                        detail = response.json()

                    except:

                        detail = response.text

                    

                    if status_code == 404:

                        error_msg = f"路径错误 (404)。请确保请求地址正确：{url}"

                    elif status_code == 401:

                        error_msg = "鉴权失败 (401)。请检查 API Key 权限。"

                    else:

                        error_msg = f"API 错误: {status_code} - {detail}"

                    

                    # 404 和 401 通常是配置问题，不进行盲目重试

                    if status_code in [404, 401]: break

                    response.raise_for_status()

                # 处理 SSE 流

                for line in response.iter_lines():

                    if not line: continue

                    line_text = line.decode('utf-8').strip()

                    

                    # 剥离 "data: " 前缀

                    json_str = line_text[5:].strip() if line_text.startswith("data:") else line_text

                    if json_str == "[DONE]": break

                    

                    try:

                        chunk = json.loads(json_str)

                        delta_text = ""

                        # 兼容 OpenAI/DeepSeek/Gemini-OpenAI-Adapter 格式

                        if "choices" in chunk and len(chunk["choices"]) > 0:

                            choice = chunk["choices"][0]

                            if "delta" in choice:

                                delta_text = choice["delta"].get("content", "")

                            elif "message" in choice:

                                delta_text = choice["message"].get("content", "")

                        # 兼容 Qwen 格式

                        elif "output" in chunk:

                            output = chunk["output"]

                            if "choices" in output:

                                delta_text = output["choices"][0].get("message", {}).get("content", "")

                            elif "text" in output:

                                delta_text = output["text"]

                        

                        if delta_text:

                            full_content += delta_text

                    except json.JSONDecodeError:

                        continue

            if full_content:

                streaming_placeholder.empty()

                return True, {"choices": [{"message": {"content": full_content}}]}, ""

            else:

                error_msg = "模型未返回有效文本内容。"

        except Exception as e:

            error_msg = f"请求异常（第{attempt+1}次尝试）: {str(e)}"

            time.sleep(2 ** attempt)

    streaming_placeholder.empty()

    return False, {"error": error_msg}, error_msg

# ===============================
# 词类判定主函数
# ===============================

def ask_model_for_pos_and_scores(word: str, provider: str, model: str, api_key: str) -> Tuple[Dict[str, Dict[str, int]], str, str, str]:

    """词类判定核心函数"""

    if not word:

        return {}, "", "未知", ""

    full_rules_by_pos = {

        pos: "\n".join([f"- {r['name']}: {r['desc']}（符合: {r['match_score']} 分，不符合: {r['mismatch_score']} 分）" for r in rules])

        for pos, rules in RULE_SETS.items()

    }

    system_msg = f"""你是一名中文词法与语法方面的专家。现在要分析词语「{word}」在下列词类中的表现：

- 需要判断的词类：名词、动词、名动词

- 评分规则已经由系统定义，你**不要**自己设计分值，也**不要**在 JSON 中给出具体数字分数。程序将根据你的判断（true/false）自动赋值。

- 你只需要判断每一条规则是"符合"还是"不符合"。

【各词类的规则说明（仅供你判断使用）】

【名词】

{full_rules_by_pos["名词"]}

【动词】

{full_rules_by_pos["动词"]}

【名动词】

{full_rules_by_pos["名动词"]}

【输出要求】

1. 在 explanation 字段中，必须**逐条规则**说明判断依据，并举例（可以自己造句）：

   - 格式示例：

     - 「名词-N1_可受数量词修饰：符合。理由：……。例句：……。」

     - 「动词-V2_可后附/插入时体助词'着/了/过'：不符合。理由：……。例句：……。」

   - explanation 里要覆盖 **三个词类的所有规则**，不能只写几条。

2. 在 JSON 中的 scores 字段里：

   - 每一类下的每一条规则，只能给出 **布尔值 true / false**，表示是否符合该规则

   - 严禁在 scores 里使用数值分数（例如 0, 5, 10 等）

   - 如果你不确定，也必须做出判断（true 或 false），不要用 null、0 或其它值

   - JSON 结构必须是：{{"explanation": "...", "predicted_pos": "...", "scores": {{"名词": {{...}}, "动词": {{...}}, "名动词": {{...}}}}}}

3. predicted_pos：

   - 请选择「名词」「动词」「名动词」之一，作为该词语最典型的词类。

4. **最后输出时，先写详细的文字推理，最后单独且完整地给出一段合法的 JSON（不要再加注释）。**

"""

    user_prompt = f"""

请严格按照上述要求分析词语「{word}」。

特别注意：

- 在 JSON 的 scores 部分，只能用 true/false 表示"是否符合规则"，不能使用任何数字。

- explanation 中必须对每一条规则写明"符合/不符合 + 理由 + 例句"。

请先给出详细推理过程，然后在最后单独输出一个 JSON 对象。

"""

    with st.spinner(f"正在调用大模型 ({model}) 进行分析，请稍候..."):

        ok, resp_json, err_msg = call_llm_api_cached(

            _provider=provider,

            _model=model,

            _api_key=api_key,

            messages=[

                {"role": "system", "content": system_msg},

                {"role": "user", "content": user_prompt}

            ]

        )

    if not ok:

        st.error(f"模型调用失败: {err_msg}")

        logger.error(f"模型调用失败 - 词语:{word}, 错误:{err_msg}")

        return {}, f"调用失败: {err_msg}", "未知", f"模型调用失败: {err_msg}"

    raw_text = extract_text_from_response(resp_json)

    parsed_json, cleaned_json_text = extract_json_from_text(raw_text)

    if parsed_json and isinstance(parsed_json, dict):

        explanation = parsed_json.get("explanation", "模型未提供详细推理过程。")

        predicted_pos = parsed_json.get("predicted_pos", "未知")

        raw_scores = parsed_json.get("scores", {})

        if predicted_pos not in RULE_SETS:

             st.warning(f"模型预测的词类 '{predicted_pos}' 不在分析范围内 ('名词', '动词', '名动词')。")

    else:

        st.error(" 未能从模型响应中解析出有效的JSON。请检查模型输出是否符合要求。")

        explanation = "无法解析模型输出。原始响应：\n" + raw_text

        predicted_pos = "未知"

        raw_scores = {}

        cleaned_json_text = raw_text

    scores_out = {pos: {} for pos in RULE_SETS.keys()}

    try:

        for pos, rules in RULE_SETS.items():

            raw_pos_scores = raw_scores.get(pos, {})

            if isinstance(raw_pos_scores, dict):

                for k, v in raw_pos_scores.items():

                    normalized_key = normalize_key(k, rules)

                    if normalized_key:

                        rule_def = next(r for r in rules if r["name"] == normalized_key)

                        scores_out[pos][normalized_key] = map_to_allowed_score(rule_def, v)

        # 补全缺失的规则得分

        for pos, rules in RULE_SETS.items():

            for rule in rules:

                rule_name = rule["name"]

                if rule_name not in scores_out[pos]:

                    scores_out[pos][rule_name] = 0

    except Exception as e:

        logger.error(f"处理得分失败: {e}")

        scores_out = {}

    return scores_out, raw_text, predicted_pos, explanation

# ===============================
# 雷达图绘制函数
# ===============================

def plot_radar_chart_streamlit(scores_norm: Dict[str, float], title: str):

    """绘制词类隶属度雷达图"""

    if not scores_norm:

        st.warning("无法绘制雷达图：没有有效数据。")

        return

    

    categories = list(scores_norm.keys())

    if not categories:

        st.warning("无法绘制雷达图：没有有效词类。")

        return

        

    values = list(scores_norm.values())

    categories += [categories[0]]

    values += [values[0]]

    

    min_val = min(values)

    max_val = max(values)

    axis_min = min(min_val, -0.1) 

    axis_max = max(max_val, 1.0)

    

    fig = go.Figure(data=[

        go.Scatterpolar(

            r=values, 

            theta=categories, 

            fill="toself", 

            name="隶属度",

            hovertemplate = '<b>%{theta}</b><br>隶属度: %{r:.4f}<extra></extra>'

        )

    ])

    fig.update_layout(

        polar=dict(

            radialaxis=dict(

                visible=True, 

                range=[axis_min, axis_max],

                tickvals=[0, 0.25, 0.5, 0.75, 1.0] if axis_min >= 0 else [-1.0, -0.5, 0, 0.5, 1.0]

            )

        ),

        showlegend=False,

        title=dict(text=title, x=0.5, font=dict(size=16))

    )

    st.plotly_chart(fig, use_container_width=True)

# ===============================
# 增强型批量处理（核心修复中断）
# ===============================

def process_and_style_excel(df, selected_model_info, target_col_name, metric_placeholder, backup_file):

    """批量处理Excel并实时更新数据量，增强鲁棒性"""

    output = io.BytesIO()

    if 'processed_history' not in st.session_state:

        st.session_state.processed_history = []

    

    progress_bar = st.progress(0)

    status_text = st.empty()

    backup_info_placeholder = st.container()

    total = len(df)

    file_name = f"excel_{int(time.time())}"  # 唯一标识当前文件

    

    # 加载上次进度

    last_progress = load_process_progress()

    start_row = 0

    if last_progress and last_progress.get("file_name") == file_name:

        start_row = last_progress.get("current_row", 0)

        st.info(f"检测到上次未完成的任务，从第 {start_row+1} 行继续处理")

    

    try:

        for index in range(start_row, total):

            row = df.iloc[index]

            word = str(row[target_col_name]).strip()

            

            # 保存当前进度

            save_process_progress(file_name, index, total)

            

            # 单条数据异常捕获（核心：避免单条失败中断整体）

            try:

                max_retries = 3

                success = False

                scores_all, raw_text, predicted_pos, explanation = {}, "", "请求失败", ""

                

                for attempt in range(max_retries):

                    try:

                        status_text.text(f"正在处理 ({index + 1}/{total}): {word} ... (尝试 {attempt + 1})")

                        scores_all, raw_text, predicted_pos, explanation = ask_model_for_pos_and_scores(

                            word=word,

                            provider=selected_model_info["provider"],

                            model=selected_model_info["model"],

                            api_key=selected_model_info["api_key"]

                        )

                        if scores_all:

                            success = True

                            break

                        time.sleep(2)

                    except Exception as e:

                        logger.error(f"处理词语{word}失败（尝试{attempt+1}）: {e}")

                        time.sleep(2)

                

                # 构造数据行

                membership = calculate_membership(scores_all) if success else {}

                new_row = {

                    "序数": index + 1,

                    "词语": word,

                    "动词": membership.get("动词", 0.0),

                    "名词": membership.get("名词", 0.0),

                    "名动词": membership.get("名动词", 0.0),

                    "差值/距离": round(abs(membership.get("动词", 0.0) - membership.get("名词", 0.0)), 4),

                    "预测词类": predicted_pos,

                    "原始响应": raw_text if success else f"错误: {explanation}",

                    "时间戳": time.strftime("%Y-%m-%d %H:%M:%S")

                }

                

                # 保存到SessionState

                st.session_state.processed_history.append(new_row)

                

                # 安全写入CSV并实时更新数据量

                try:

                    temp_df = pd.DataFrame([new_row])

                    header_needed = not os.path.exists(backup_file)

                    write_success = safe_write_csv(temp_df, backup_file, mode='a', header=header_needed)

                    if write_success:

                        # 实时更新已存数据量

                        latest_count = get_history_count(backup_file)

                        metric_placeholder.metric("已存数据量", f"{latest_count} 条")

                    else:

                        st.error(f" 保存第 {index+1} 条记录失败（文件写入错误）")

                except Exception as csv_err:

                    st.error(f"保存第 {index+1} 条记录失败: {csv_err}")

                    logger.error(f"保存CSV失败 - 行号:{index+1}, 错误:{csv_err}")

                with backup_info_placeholder:

                    st.info(f"已自动保存第 {index+1} 条记录。如遇中断，下次将从第 {index+2} 行继续")

                progress_bar.progress((index + 1) / total)

                time.sleep(0.5)  # 限流：避免请求过快被封禁

                

            except Exception as row_err:

                logger.error(f"处理第{index+1}行失败（跳过）: {row_err}")

                st.warning(f" 跳过第 {index+1} 行（处理失败）: {row_err}")

                progress_bar.progress((index + 1) / total)

                continue

    except Exception as e:

        st.error(f"批量处理意外中断: {e}")

        logger.error(f"批量处理中断: {e}")

        return None

    finally:

        # 处理完成/中断后清除进度文件

        clear_process_progress()

    

    # 导出Excel

    final_data = st.session_state.processed_history

    if not final_data:

        return None

    result_df = pd.DataFrame(final_data)

    try:

        with pd.ExcelWriter(output, engine='openpyxl') as writer:

            cols = ["词语", "动词", "名词", "名动词", "差值/距离", "预测词类", "原始响应"]

            result_df[cols].to_excel(writer, index=False, sheet_name='分析结果')

            

            workbook = writer.book

            worksheet = writer.sheets['分析结果']

            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            

            for i, data_row in enumerate(final_data):

                row_num = i + 2

                pred = data_row["预测词类"]

                target_idx = {"动词": 2, "名词": 3, "名动词": 4}.get(pred)

                if target_idx:

                    worksheet.cell(row=row_num, column=target_idx).fill = yellow_fill

                    

        return output.getvalue()

    except Exception as e:

        st.error(f"Excel 生成失败: {e}")

        logger.error(f"生成Excel失败: {e}")

        return None

# ===============================
# 主页面逻辑（UI优化版）
# ===============================

def main():
    # ===== 顶部标题高亮卡片 =====
    st.markdown("""
    <div class="title-header-card">
        <h1>基于大语言模型的汉语词类隶属度检测划类平台</h1>
        <div class="subtitle">Chinese Membership Detection and Classification Platform Based on Large Language Models (LLMs)</div>
        <div class="badges">
            <span class="badge">多模型支持</span>
            <span class="badge">隶属度分析</span>
            <span class="badge">可视化展示</span>
            <span class="badge">批量处理</span>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ===== 顶部控制区 =====
    control_container = st.container()
    with control_container:
        col1, col2 = st.columns([3, 1])
        
        with col1:
            # 模型设置高亮容器
            st.markdown('<div class="model-settings-card">', unsafe_allow_html=True)
            st.markdown('<div class="section-title"><span class="icon-dot"></span> 模型设置（LLM）</div>', unsafe_allow_html=True)
            
            if not AVAILABLE_MODEL_OPTIONS:
                st.markdown('<div class="error-highlight">', unsafe_allow_html=True)
                st.error("找不到可用的 API Key！请设置以下任意一个环境变量来启用模型:")
                for name, info in MODEL_OPTIONS.items():
                    st.code(f"export {info['env_var']}='你的API Key'", language="bash")
                st.markdown('</div>', unsafe_allow_html=True)
                selected_model_display_name = list(MODEL_OPTIONS.keys())[0]
                selected_model_info = MODEL_OPTIONS[selected_model_display_name]
                st.selectbox("选择大模型 (不可用)", list(MODEL_OPTIONS.keys()), disabled=True)
            else:
                selected_model_display_name = st.selectbox(
                    "选择大模型", 
                    list(AVAILABLE_MODEL_OPTIONS.keys()), 
                    key="model_select"
                )
                selected_model_info = AVAILABLE_MODEL_OPTIONS[selected_model_display_name]
                # 显示当前模型状态
                st.markdown(f"""
                <div style="display: flex; align-items: center; gap: 0.5rem; margin-top: 0.5rem;">
                    <span class="status-badge success">● 已配置</span>
                    <span style="color: #64748b; font-size: 0.85rem;">提供商: {selected_model_info['provider'].upper()}</span>
                </div>
                """, unsafe_allow_html=True)
            
            st.markdown('</div>', unsafe_allow_html=True)
                
        with col2:
            # 连接测试高亮容器
            st.markdown('<div class="connection-card">', unsafe_allow_html=True)
            st.markdown('<div class="section-title" style="justify-content: center;"><span class="icon-dot"></span> 连接测试</div>', unsafe_allow_html=True)
            st.write("")
            if not selected_model_info["api_key"]:
                st.button("测试模型链接 (不可用)", type="secondary", disabled=True)
            else:
                if st.button("测试模型链接", type="secondary"):
                    with st.spinner("正在测试连接..."):
                        ok, _, err_msg = call_llm_api_cached(
                            _provider=selected_model_info["provider"],
                            _model=selected_model_info["model"],
                            _api_key=selected_model_info["api_key"],
                            messages=[{"role": "user", "content": "请回复'pong'"}],
                            max_tokens=10
                        )
                    if ok:
                        st.success("成功！")
                    else:
                        st.error(f"失败: {err_msg}")
            st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("---")

    # ===== 分页 =====
    tab1, tab2 = st.tabs(["单个词语详细分析", "Excel 批量处理"])

    # ===== 单个词语分析 =====
    with tab1:
        # 输入区卡片
        st.markdown('<div class="module-card">', unsafe_allow_html=True)
        st.markdown('<div class="section-title"><span class="icon-dot"></span> 词语输入</div>', unsafe_allow_html=True)
        word = st.text_input("请输入要分析的汉语词语", placeholder="例如：苹果、跑、美丽...", key="word_input")
        analyze_button = st.button(
            "开始分析", 
            type="primary",
            disabled=not (selected_model_info["api_key"] and word)
        )
        
        with st.expander("ℹ️ 使用说明", expanded=False):
            st.markdown('<div class="info-highlight">', unsafe_allow_html=True)
            st.info("""
            1. **配置 API Key**: 请在运行程序前设置必要的环境变量。
            2. **词语输入**：在上方的"词语输入"框中输入一个汉语词。
            3. **开始分析**：点击"开始分析"按钮。
            4. **结果解析**：系统将显示隶属度、雷达图和详细规则得分。
            """)
            st.markdown('</div>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

        if analyze_button and word and selected_model_info["api_key"]:
            status_placeholder = st.empty()
            status_placeholder.info(f"正在为词语「{word}」启动分析，使用模型：{selected_model_display_name}...")

            scores_all, raw_text, predicted_pos, explanation = ask_model_for_pos_and_scores(
                word=word,
                provider=selected_model_info["provider"],
                model=selected_model_info["model"],
                api_key=selected_model_info["api_key"]
            )
            
            status_placeholder.empty()
            
            if scores_all:
                membership = calculate_membership(scores_all)
                final_membership = membership.get(predicted_pos, 0)
                
                # 成功结果高亮卡片
                st.markdown(f"""
                <div class="result-success-card">
                    <div style="font-size: 1.1rem; font-weight: 600; color: #065f46;">
                        分析完成
                    </div>
                    <div style="margin-top: 0.5rem; font-size: 1rem; color: #065f46;">
                        词语「<strong>{word}</strong>」最可能的词类是 
                        <span style="background: #10b981; color: white; padding: 0.2rem 0.6rem; border-radius: 6px; font-weight: 600;">{predicted_pos}</span>
                        ，隶属度为 <strong>{final_membership:.4f}</strong>
                    </div>
                </div>
                """, unsafe_allow_html=True)
                
                col_results_1, col_results_2 = st.columns(2)
                
                with col_results_1:
                    # 隶属度排名卡片
                    st.markdown('<div class="module-card">', unsafe_allow_html=True)
                    st.markdown('<div class="section-title"><span class="icon-dot"></span> 词类隶属度排名</div>', unsafe_allow_html=True)
                    top10 = get_top_10_positions(membership)
                    
                    # 自定义排名卡片展示
                    for i, (pos, score) in enumerate(top10):
                        rank_class = f"top-{i+1}" if i < 3 else ""
                        st.markdown(f"""
                        <div class="rank-card {rank_class}">
                            <div style="display: flex; align-items: center; gap: 0.75rem;">
                                <div class="rank-num">{i+1}</div>
                                <span style="font-weight: 600; color: #1e3a5f;">{pos}</span>
                            </div>
                            <span style="font-weight: 700; color: #2d5a87; font-size: 1.1rem;">{score:.4f}</span>
                        </div>
                        """, unsafe_allow_html=True)
                    st.markdown('</div>', unsafe_allow_html=True)
                    
                    # 雷达图卡片
                    st.markdown('<div class="module-card">', unsafe_allow_html=True)
                    st.markdown('<div class="section-title"><span class="icon-dot"></span> 词类隶属度雷达图</div>', unsafe_allow_html=True)
                    plot_radar_chart_streamlit(dict(top10), f"「{word}」的词类隶属度分布")
                    st.markdown('</div>', unsafe_allow_html=True)

                with col_results_2:
                    # 详细得分卡片
                    st.markdown('<div class="module-card">', unsafe_allow_html=True)
                    st.markdown('<div class="section-title"><span class="icon-dot"></span> 各词类详细得分</div>', unsafe_allow_html=True)
                    pos_total_scores = {pos: sum(scores_all[pos].values()) for pos in scores_all.keys()}
                    sorted_pos_names = sorted(pos_total_scores.keys(), key=lambda pos: pos_total_scores[pos], reverse=True)
                    
                    for pos in sorted_pos_names:
                        total_score = pos_total_scores[pos]
                        max_rule = max(scores_all[pos].items(), key=lambda x: x[1], default=("无", 0))
                        with st.expander(f"**{pos}** (总分: {total_score}, 最高分规则: {max_rule[0]} - {max_rule[1]}分)"):
                            rule_data = []
                            for rule_name, rule_score in scores_all[pos].items():
                                rule_desc = ""
                                if pos in RULE_SETS:
                                    for rule in RULE_SETS[pos]:
                                        if rule["name"] == rule_name:
                                            rule_desc = rule["desc"]
                                            break
                                rule_data.append({
                                    "规则代码": rule_name,
                                    "规则描述": rule_desc,
                                    "得分": rule_score
                                })
                            rule_data_sorted = sorted(rule_data, key=lambda x: x["得分"], reverse=True)
                            rule_df = pd.DataFrame(rule_data_sorted)
                            styled_df = rule_df.style.map(
                                lambda x: "color: #ff4b4b; font-weight: bold"
                                if isinstance(x, (int, float)) and x < 0 else "",
                                subset=["得分"]
                            )
                            st.dataframe(
                                styled_df,
                                use_container_width=True,
                                height=min(len(rule_df) * 30 + 50, 400)
                            )
                    st.markdown('</div>', unsafe_allow_html=True)
                    
                    # 原始响应卡片
                    st.markdown('<div class="module-card">', unsafe_allow_html=True)
                    st.markdown('<div class="section-title"><span class="icon-dot"></span> 模型原始响应</div>', unsafe_allow_html=True)
                    with st.expander("点击展开查看原始响应", expanded=False):
                        st.code(raw_text, language="text")
                    st.markdown('</div>', unsafe_allow_html=True)

    # ===== 批量处理 =====
    with tab2:
        # 批量任务监控标题
        st.markdown('<div class="module-card">', unsafe_allow_html=True)
        st.markdown('<div class="section-title"><span class="icon-dot"></span> 批量任务实时监控</div>', unsafe_allow_html=True)
        
        # 控制面板
        st.markdown("#### 控制面板")
        ctrl_col1, ctrl_col2, ctrl_col3 = st.columns([2, 1, 1])
        
        with ctrl_col1:
            # 可实时更新的metric占位符
            metric_placeholder = st.empty()
            # 初始化显示最新数量
            history_count = get_history_count(BACKUP_FILE)
            metric_placeholder.metric("已存数据量", f"{history_count} 条")
            
            has_history = os.path.exists(BACKUP_FILE)
            if has_history:
                st.caption(f"存储位置: `{BACKUP_FILE}`")
        
        with ctrl_col2:
            if os.path.exists(BACKUP_FILE):
                with open(BACKUP_FILE, "rb") as f:
                    st.download_button(
                        label="下载历史文件(CSV)",
                        data=f,
                        file_name=f"batch_results_{time.strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv",
                        use_container_width=True
                    )
            else:
                st.button("下载历史文件", disabled=True, use_container_width=True)
        with ctrl_col3:
            if st.button("清空本地记录", use_container_width=True, type="secondary"):
                if os.path.exists(BACKUP_FILE):
                    try:
                        os.remove(BACKUP_FILE)
                        clear_process_progress()  # 同时清除进度
                        st.success("已清空本地记录和进度")
                        metric_placeholder.metric("已存数据量", "0 条")
                        st.rerun()
                    except Exception as e:
                        st.error(f"清空记录失败: {e}")
                else:
                    st.info("暂无本地记录可清空")
        
        st.divider()
        
        # 运行状态
        st.markdown("#### 运行状态")
        progress_bar = st.progress(0)
        status_info = st.empty()
        
        # 实时结果预览
        st.markdown("#### 实时结果预览")
        table_placeholder = st.empty()
        if os.path.exists(BACKUP_FILE):
            try:
                table_placeholder.dataframe(
                    pd.read_csv(BACKUP_FILE, encoding='utf-8-sig'), 
                    use_container_width=True, 
                    height=300
                )
            except Exception as e:
                table_placeholder.error(f"显示历史记录失败: {e}")
        else:
            table_placeholder.info("暂无数据。上传文件并点击开始后，结果将在此逐行实时显示。")
        
        st.divider()
        
        # 上传任务
        st.markdown("#### 上传新任务")
        uploaded_file = st.file_uploader("选择 Excel 文件", type=["xlsx", "xls"])
        
        if uploaded_file:
            try:
                df_input = pd.read_excel(uploaded_file)
                target_col = next((col for col in df_input.columns if "词" in str(col) or "word" in str(col).lower()), None)
                
                if target_col:
                    st.markdown(f"""
                    <div class="info-highlight">
                        <div style="font-weight: 600; color: #1e40af;">文件信息</div>
                        <div style="margin-top: 0.5rem;">
                            识别到目标列: <code>{target_col}</code> | 待分析总数: <strong>{len(df_input)}</strong> 条
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    if st.button("开始处理", type="primary", use_container_width=True):
                        if not selected_model_info["api_key"]:
                            st.error("请先在上方配置有效的 API Key")
                        else:
                            # 获取已处理的词语
                            existing_words = set()
                            if os.path.exists(BACKUP_FILE):
                                try:
                                    existing_df = pd.read_csv(BACKUP_FILE, encoding='utf-8-sig')
                                    if "词语" in existing_df.columns:
                                        existing_words = set(existing_df["词语"].astype(str).tolist())
                                    st.info(f"已跳过 {len(existing_words)} 条已处理记录")
                                except Exception as e:
                                    st.warning(f"读取已处理记录失败，将重新处理所有数据: {e}")
                            total_rows = len(df_input)
                            
                            # 批量处理主循环（全量异常捕获）
                            try:
                                for index, row in df_input.iterrows():
                                    word = str(row[target_col]).strip()
                                    if not word:
                                        status_info.write(f"**跳过空值**: 第 {index+1}/{total_rows} 行")
                                        progress_bar.progress((index + 1) / total_rows)
                                        continue
                                    
                                    pct = int((index + 1) / total_rows * 100)
                                    progress_bar.progress((index + 1) / total_rows)
                                    
                                    if word in existing_words:
                                        status_info.write(f" **跳过已处理**: {word} ({index+1}/{total_rows}) | 进度: {pct}%")
                                        continue
                                    
                                    status_info.write(f" **正在分析**: `{word}` | 进度: {index+1}/{total_rows} ({pct}%)")
                                    
                                    # 调用API处理（增强重试）
                                    max_retries = 3
                                    success = False
                                    scores, raw_text, pred_pos, explanation = {}, "", "处理失败", "无响应"
                                    for attempt in range(max_retries):
                                        try:
                                            scores, raw_text, pred_pos, explanation = ask_model_for_pos_and_scores(
                                                word=word,
                                                provider=selected_model_info["provider"],
                                                model=selected_model_info["model"],
                                                api_key=selected_model_info["api_key"]
                                            )
                                            success = bool(scores)
                                            if success:
                                                break
                                            time.sleep(2)
                                        except Exception as e:
                                            explanation = f"调用异常: {str(e)}"
                                            logger.error(f"处理词语{word}失败（尝试{attempt+1}）: {e}")
                                            time.sleep(2)
                                    
                                    # 构造数据行
                                    membership = calculate_membership(scores) if success else {}
                                    new_row = {
                                        "序数": index + 1,
                                        "词语": word,
                                        "动词": membership.get("动词", 0.0),
                                        "名词": membership.get("名词", 0.0),
                                        "名动词": membership.get("名动词", 0.0),
                                        "差值/距离": round(abs(membership.get("动词", 0.0) - membership.get("名词", 0.0)), 4),
                                        "预测词类": pred_pos,
                                        "原始响应": raw_text if success else f"错误: {explanation}",
                                        "时间戳": time.strftime("%Y-%m-%d %H:%M:%S")
                                    }
                                    
                                    # 安全保存数据
                                    try:
                                        temp_df = pd.DataFrame([new_row])
                                        header_needed = not os.path.exists(BACKUP_FILE)
                                        write_success = safe_write_csv(temp_df, BACKUP_FILE, mode='a', header=header_needed)
                                        if write_success:
                                            existing_words.add(word)
                                            # 实时更新已存数据量
                                            latest_count = get_history_count(BACKUP_FILE)
                                            metric_placeholder.metric("已存数据量", f"{latest_count} 条")
                                        else:
                                            st.error(f"保存第 {index+1} 条记录失败（文件写入错误）")
                                    except Exception as csv_err:
                                        st.error(f"保存第 {index+1} 条记录失败: {csv_err}")
                                        logger.error(f"保存CSV失败 - 行号:{index+1}, 错误:{csv_err}")
                                    
                                    # 刷新表格
                                    try:
                                        updated_df = pd.read_csv(BACKUP_FILE, encoding='utf-8-sig')
                                        table_placeholder.dataframe(updated_df, use_container_width=True, height=300)
                                    except Exception as read_err:
                                        st.warning(f"刷新表格失败: {read_err}")
                                    
                                    time.sleep(0.5)  # 限流
                                
                                progress_bar.progress(100)
                                status_info.success(f"批量处理完成！总处理量: {total_rows} 条，已保存到 {BACKUP_FILE}")
                                clear_process_progress()  # 清除进度
                                st.rerun()
                            except Exception as batch_err:
                                logger.error(f"批量处理主循环中断: {batch_err}")
                                status_info.error(f" 批量处理中断: {batch_err}，下次可从断点继续")
                else:
                    st.markdown('<div class="error-highlight">', unsafe_allow_html=True)
                    st.error("未识别到包含'词'或'word'的列，请检查Excel文件结构")
                    st.markdown('</div>', unsafe_allow_html=True)
            except Exception as e:
                st.error(f"读取Excel文件失败: {e}")
                logger.error(f"读取Excel失败: {e}")
        
        st.markdown('</div>', unsafe_allow_html=True)

# ===============================
# 运行主函数
# ===============================

if __name__ == "__main__":

    main()

# ===============================
# 页面底部说明
# ===============================
st.markdown("---")
st.markdown(
    """
    <div class="footer-text">
        © 2025 汉语词类隶属度检测划类平台 | Ryan
    </div>
    """,
    unsafe_allow_html=True
)
